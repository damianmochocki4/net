from tkinter import Tk, filedialog, messagebox
from pandas import DataFrame, ExcelWriter
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Border, Side
from os import path, walk, remove

# Function to select the directory
def select_directory():
    root = Tk()
    root.withdraw()  # Close the root window
    folder_selected = filedialog.askdirectory()
    return folder_selected

# Function to find the first folder starting with "Re" in the selected directory path
def find_re_folder(path):
    parts = path.replace("\\", "/").split("/")
    for part in parts:
        if part.startswith("Re"):
            return part
    return "Inventory"  # Default name if no folder starting with 'Re' is found

# Function to scan a folder and create Excel summary if it contains files
def process_folder(folder_path, aggregated_data, deleted_files_data):
    folder_name = path.basename(folder_path)
    summary_data = []
    local_deleted_files_data = []  # Only for the current folder
    
    subdirs = next(walk(folder_path))[1]
    files = next(walk(folder_path))[2]
    
    num_subfolders = len(subdirs)
    num_files = len(files)
    num_pdf_files = len([file for file in files if file.lower().endswith('.pdf')])
    num_jpeg_files = len([file for file in files if file.lower().endswith(('.jpg', '.jpeg'))])
    num_other_files = num_files - num_pdf_files - num_jpeg_files
    num_deleted_files = 0

    for file in files:
        if file.lower().endswith('.pdf'):
            file_path = path.join(folder_path, file)
            file_size = path.getsize(file_path)
            delete_time = datetime.now().strftime("%d.%m.%Y %H:%M")
            
            deleted_files_data.append({
                "Dateiname": file,
                "Ordnername": folder_name,
                "Dateityp": "PDF",
                "Löschdatum": delete_time,
                "Dateigröße": f"{file_size / 1024:.0f}kB",
            })
            
            local_deleted_files_data.append({
                "Dateiname": file,
                "Ordnername": folder_name,
                "Dateityp": "PDF",
                "Löschdatum": delete_time,
                "Dateigröße": f"{file_size / 1024:.0f}kB",
            })
            
            remove(file_path)  # Delete the file
            num_deleted_files += 1

    # Add headers if no PDF files were deleted
    if num_pdf_files == 0:
        local_deleted_files_data.append({
            "Dateiname": "N/A",
            "Ordnername": folder_name,
            "Dateityp": "N/A",
            "Löschdatum": "N/A",
            "Dateigröße": "N/A",
        })
    
    if num_files > 0:
        summary_row = {
            "Ordnername": folder_name,
            "Number of subfolders": num_subfolders,
            "Number of files": num_files,
            "Number of PDF files": num_pdf_files,
            "Number of JPEG files": num_jpeg_files,
            "Number of other files": num_other_files,
            "Number of deleted files": num_deleted_files
        }
        summary_data.append(summary_row)
        aggregated_data.append(summary_row)
    
    if summary_data:
        save_to_excel(folder_path, folder_name, summary_data, local_deleted_files_data)

# Function to save data into Excel for each folder and adjust column width
def save_to_excel(folder_path, folder_name, summary_data, deleted_files_data):
    excel_filename = f"Register_{folder_name}.xlsx"
    excel_path = path.join(folder_path, excel_filename)
    
    with ExcelWriter(excel_path, engine='openpyxl') as writer:
        summary_df = DataFrame(summary_data)
        summary_df.to_excel(writer, sheet_name='Folder Summary', index=False)
    
    wb = load_workbook(excel_path)
    sheet = wb['Folder Summary']
    start_row = sheet.max_row + 2
    sheet[f"A{start_row}"] = "Gelöschte Dateien / Deleted files"
    sheet[f"A{start_row}"].font = Font(bold=True)
    
    # Add headers with styling
    header_row = start_row + 1
    headers = ["Dateiname", "Ordnername", "Dateityp", "Löschdatum", "Dateigröße"]
    for col_num, header in enumerate(headers, start=1):
        cell = sheet.cell(row=header_row, column=col_num, value=header)
        cell.font = Font(bold=True)

        # Apply borders only to header cells
        cell.border = Border(left=Side(style='thin'), 
                             right=Side(style='thin'), 
                             top=Side(style='thin'), 
                             bottom=Side(style='thin'))

    if deleted_files_data:
        deleted_files_df = DataFrame(deleted_files_data, columns=headers)
        
        for r_idx, row in enumerate(dataframe_to_rows(deleted_files_df, index=False, header=False), start=header_row + 1):
            for c_idx, value in enumerate(row, start=1):
                cell = sheet.cell(row=r_idx, column=c_idx, value=value)
                # No border for data cells

    adjust_column_width(sheet)
    wb.save(excel_path)
    print(f"Excel file '{excel_filename}' created in '{folder_path}'")
    
# Function to adjust column width in each sheet
def adjust_column_width(sheet):
    for col in sheet.columns:
        max_length = 0
        column = col[0].column_letter
        
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        
        adjusted_width = max_length + 2
        sheet.column_dimensions[column].width = adjusted_width

# Function to save the aggregated summary in a separate Excel file
def save_aggregated_summary(base_path, folder_name, aggregated_data, deleted_files_data, excel_subfolders_count):
    summary_df = DataFrame(aggregated_data)
    
    summary_filename = "Inventory.xlsx" if folder_name == "Inventory" else f"Inventory_{folder_name}.xlsx"
    parent_path = path.dirname(base_path)
    
    summary_path = path.normpath(path.join(parent_path, summary_filename))
    
    with ExcelWriter(summary_path, engine='openpyxl') as writer:
        # Create and save the first summary table
        structured_summary_df = DataFrame({
            "Invoice name": [folder_name],
            "Number of subfolders": [excel_subfolders_count],
            "Number of files before deletion incl. Inventory file": [summary_df["Number of files"].sum() + 1],
            "Number of PDF files": [summary_df["Number of PDF files"].sum()],
            "Number of JPEG files": [summary_df["Number of JPEG files"].sum()],
            "Number of other files": [summary_df["Number of other files"].sum()],
            "Number of deleted files": [summary_df["Number of deleted files"].sum()],
        })

        structured_summary_df.to_excel(writer, sheet_name='Aggregated Summary', index=False)

        # Adding header styling and borders for the top summary
        header_row = 1  # Start from the first row
        headers = structured_summary_df.columns.tolist()

        for col_num, header in enumerate(headers, start=1):
            cell = writer.sheets['Aggregated Summary'].cell(row=header_row, column=col_num, value=header)
            cell.font = Font(bold=True)

            # Apply borders only to header cells
            cell.border = Border(left=Side(style='thin'), 
                                 right=Side(style='thin'), 
                                 top=Side(style='thin'), 
                                 bottom=Side(style='thin'))

        # Adding space before the deleted files section
        start_row_for_deleted_files = header_row + 3  # Leave two rows of space
        
        if deleted_files_data:
            deleted_files_df = DataFrame(deleted_files_data, columns=[
                "Dateiname", "Ordnername", "Dateityp", "Löschdatum", "Dateigröße"
            ])
            
            # Write the headers for the deleted files section
            deleted_files_headers = ["Dateiname", "Ordnername", "Dateityp", "Löschdatum", "Dateigröße"]
            for col_num, header in enumerate(deleted_files_headers, start=1):
                cell = writer.sheets['Aggregated Summary'].cell(row=start_row_for_deleted_files, column=col_num, value=header)
                cell.font = Font(bold=True)
                cell.border = Border(left=Side(style='thin'), 
                                     right=Side(style='thin'), 
                                     top=Side(style='thin'), 
                                     bottom=Side(style='thin'))

            # Write the deleted files data
            for r_idx, row in enumerate(dataframe_to_rows(deleted_files_df, index=False, header=False), start=start_row_for_deleted_files + 1):
                for c_idx, value in enumerate(row, start=1):
                    cell = writer.sheets['Aggregated Summary'].cell(row=r_idx, column=c_idx, value=value)

        adjust_column_width(writer.sheets['Aggregated Summary'])

    print(f"Aggregated summary saved as '{summary_filename}' in '{parent_path}'")
    return summary_path


# Main execution: scan through each folder in the selected directory
def main():
    try:
        selected_directory = select_directory()
        if not selected_directory:
            print("No directory selected. Exiting.")
            return

        base_folder = find_re_folder(selected_directory)
        if not base_folder:
            print(f"No folder starting with 'Re' found in '{selected_directory}'. Exiting.")
            return
        
        aggregated_data = []
        deleted_files_data = []
        excel_subfolders_count = 0

        for root, subdirs, files in walk(selected_directory):
            if len(files) > 0:
                process_folder(root, aggregated_data, deleted_files_data)
                excel_subfolders_count += 1

        summary_path = save_aggregated_summary(selected_directory, base_folder, aggregated_data, deleted_files_data, excel_subfolders_count)

        # Show success message in a pop-up window
        Tk().withdraw()  # Hide main window
        messagebox.showinfo("Success", f"{excel_subfolders_count} Excels created in {excel_subfolders_count} subfolders.\n\nSummary Excel created in {summary_path}")

    except Exception as e:
        # Show error message in case of failure
        Tk().withdraw()  # Hide main window
        messagebox.showerror("Error", f"An error occurred: {str(e)}")

if __name__ == "__main__":
    main()
